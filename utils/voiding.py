from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import jpype
import asposecells
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType, PdfSaveOptions
import requests
import argparse


parser = argparse.ArgumentParser()
parser.add_argument("--back_ip", type=str)
parser.add_argument("--back_port", type=int)
parser.add_argument("--user_id", type=int)
parser.add_argument("--date", type=str)
parser.add_argument("--file_name", type=str)
parser.add_argument("--mode", type=str)
args = parser.parse_args()

CFG = {
    "back_ip":args.back_ip,
    "back_port":args.back_port,
    "user_id":args.user_id,
    "date":args.date
    }

# 기본 정보를 업데이트하는 함수
def update_basic_info(sheet, date_record,wake_time, sleep_time, wake_weight, sleep_weight):
    sheet['A1'] = date_record
    sheet['B2'] = wake_time
    sheet['B3'] = sleep_time
    sheet.merge_cells('C2:C3')
    sheet['C2'] = wake_weight
    sheet.merge_cells('D2:D3')
    sheet['D2'] = sleep_weight

    return sheet

# 배뇨 일지 정보를 추가하는 함수
def add_urination_record(sheet, index, urination_time, am_pm, urination_volume, water_intake, urgency_level):
    row = 6 + index  # 7번째 행부터 시작
    sheet[f'A{row}'] = index
    sheet[f'B{row}'] = urination_time
    if am_pm == 'AM':
        sheet[f'C{row}'] = 'V'  # 오전
    else:
        sheet[f'D{row}'] = 'V'  # 오후
    sheet[f'E{row}'] = urination_volume
    sheet[f'F{row}'] = water_intake
    urgency_columns = ['G', 'H', 'I', 'J', 'K']
    for col, level in zip(urgency_columns, urgency_level):
        sheet[f'{col}{row}'] = level
    
    return sheet

# 모든 셀을 가운데 정렬하는 함수
def center_align_cells(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    return sheet

# 모든 셀에 테두리 적용하는 함수
def apply_borders(sheet):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.border = thin_border

    return sheet

def fill_content(voiding_data, file_name):
    # 엑셀 파일 로드
    file_path = '/home/livin/kidney_service/back/doctor_page/source/배뇨일지_파이썬.xlsx'
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    
    # 데이터 입력
    sheet = update_basic_info(sheet, voiding_data['voiding']['date'],voiding_data["voiding"]["waking_time"], voiding_data["voiding"]["sleeping_time"], voiding_data["voiding"]["morning_weight"], voiding_data["voiding"]["evening_weight"])

    for n, voiding_detail in enumerate(voiding_data["voiding_details"]):
        urgency_level_list = ["","","","",""]
        urgency_level_list[voiding_detail["urgency_level"]-1] = "V"
        sheet = add_urination_record(sheet,n+1, voiding_detail["voiding_time"], 'AM', voiding_detail["voiding_volume"], voiding_detail["water_intake"], urgency_level_list)

    sheet = center_align_cells(sheet)
    sheet = apply_borders(sheet)

    new_file_path = f'./voiding_file/{file_name}.xlsx'
    wb.save(filename=new_file_path)

def get_pdf(file_name):
    file_path = f'./voiding_file/{file_name}.xlsx'
    workbook = Workbook(file_path)
    saveOptions = PdfSaveOptions()
    saveOptions.setOnePagePerSheet(True)
    workbook.save(f'./voiding_file/{file_name}.pdf', saveOptions)
    jpype.shutdownJVM()



if __name__ == "__main__":
    voiding_data = requests.get(f"http://{args.back_ip}:{args.back_port}/api/doctor/patient/voiding_details", params={"user_id":args.user_id, "date":args.date}).json()
    file_name = args.file_name
    filled_sheet = fill_content(voiding_data, file_name)
    print(f"done_{args.mode}")

    if args.mode == "pdf":
        get_pdf(file_name)